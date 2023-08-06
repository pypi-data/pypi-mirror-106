# -*- coding: utf-8 -*-
"""
封装操作 xlsx 文件功能：增、删、改、查
"""

from openpyxl import load_workbook


def check_excel(path):
    """
    显示表格所有数据，以列表形式返回。
    :param path: 文件路径。（注意：只操作第一个 Sheet1。）
    :return:
    """
    wb = load_workbook(path)
    sheet = wb.active
    for row in sheet.rows:
        cell = [row[r].value for r in range(len(row))]
        yield cell


# def update_excel(path, index, sheet='Sheet1'):
#     # 追加写
#     while True:
#         try:
#             xfile = load_workbook(path)
#             sheet = xfile.get_sheet_by_name(sheet)
#             sheet[index] = '没有该用户资料或者图片'
#             xfile.save("data.xlsx")
#             break
#         except:
#             print('保存失败，请勿打开"data.xlsx"，正在重新保存。。。')
#             time.sleep(2)
