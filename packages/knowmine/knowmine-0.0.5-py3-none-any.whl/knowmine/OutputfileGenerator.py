"""
This module contains the Output class which helps to generate
the output file of the desired format (sqlite database or excel)
containing the extracted sentences, number of sentences in the
original text (after the cleaning) and number of the extracted sentences
The result file is generated in the folder containing the articles for mining.

"""

import sqlite3
from sqlite3 import Error
import openpyxl
from openpyxl import Workbook
import os


class Output:

    def __init__(self, folder, output):
        self.dbname = 'Extracted_sentences'
        self.output = output
        self.path = folder

    def __create_connection(self):
        """
        create a database connection to the SQLite database
        specified by db_file
        :param db_file: database file
        :return: Connection object or None

        """
        conn = None
        try:
            conn = sqlite3.connect(f"{self.path+self.dbname}.db")
        except Error as e:
            print(e)

        return conn

    def __create_table(self, conn, create_table_sql):
        """
        create a table from the create_table_sql statement
        :param conn: Connection object
        :param create_table_sql: a CREATE TABLE statement
        :return:

        """
        try:
            c = conn.cursor()
            c.execute(create_table_sql)
        except Error as e:
            print(e)

    def add_result_to_database(self):

        sql_create_projects_table = """ CREATE TABLE IF NOT EXISTS texts_mining
                                            (
                                            id integer PRIMARY KEY,
                                            File_name text NOT NULL,
                                            Relevant_sentences text,
                                            N_relevant_sentences integer,
                                            N_all_sentences integer
                                            ); """

        # create a database connection
        conn = self.__create_connection()

        # create tables
        if conn is not None:
            # create projects table
            self.__create_table(conn, sql_create_projects_table)

        else:
            print("Error! cannot create the database connection.")

        sql = ''' INSERT INTO texts_mining (File_name,Relevant_sentences,
                  N_relevant_sentences,N_all_sentences)
                  VALUES(?,?,?,?)'''
        cur = conn.cursor()
        cur.execute(sql, self.output)
        conn.commit()

    def __create_excelfile(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'File_name '
        ws['B1'] = 'Relevant_sentences'
        ws['C1'] = 'N_relevant_sentences'
        ws['D1'] = 'N_all_sentences'
        wb.save(f'{self.path+self.dbname}.xlsx')

    def __write_to_excel(self):

        filename = f"{self.path+self.dbname}.xlsx"
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        column = 0
        row = ws.max_row + 1
        for info in self.output:
            column = column+1
            ws.cell(column=column, row=row, value=info)
        wb.save(filename)

    def add_result_to_excel(self):

        if os.path.exists(f'{self.path+self.dbname}.xlsx'):
            print('Existing')
            self.__write_to_excel()

        else:
            print('new')
            self.__create_excelfile()
            self.__write_to_excel()
